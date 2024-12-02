import openpyxl
from io import BytesIO

def remove_styles_from_xlsx(file):
    in_memory_file = BytesIO(file.read())
    workbook = openpyxl.load_workbook(in_memory_file)

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = openpyxl.styles.Font()
                cell.border = openpyxl.styles.Border()
                cell.fill = openpyxl.styles.PatternFill()
                cell.number_format = 'General'
                cell.protection = openpyxl.styles.Protection()
                cell.alignment = openpyxl.styles.Alignment()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output